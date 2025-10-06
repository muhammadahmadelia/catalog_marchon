import os
import sys
import json
import random
from time import sleep
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from datetime import datetime
from models.store import Store
from models.product import Product
from models.variant import Variant
from models.metafields import Metafields
import glob
import requests
from datetime import datetime
from openpyxl import Workbook
from openpyxl.drawing.image import Image as Imag
from PIL import Image

from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager

from seleniumwire import webdriver

import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)



class MyMarchon_Scraper:
    def __init__(self, DEBUG: bool, result_filename: str, logs_filename: str, chrome_path: str) -> None:
        self.DEBUG = DEBUG
        self.data = []
        self.result_filename = result_filename
        self.logs_filename = logs_filename
        self.thread_list = []
        self.thread_counter = 0
        self.chrome_options = Options()
        self.chrome_options.add_argument('--disable-infobars')
        self.chrome_options.add_argument("--start-maximized")
        self.chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])
        self.browser = webdriver.Chrome(service=ChromeService(chrome_path), options=self.chrome_options)
        pass

    def controller(self, store: Store, brands_with_types: list[dict]) -> None:
        try:
            
            self.initiate_browser(store)

            if self.login(store.username, store.password):
                if self.wait_until_element_found(20, 'xpath', '//select[@id="brands"]'):
                    
                    for brand_with_type in brands_with_types:
                        brand: str = brand_with_type['brand']
                        brand_code: str = str(brand_with_type['code']).strip().upper()
                        
                        print(f'Brand: {brand}')
                        self.print_logs(f'Brand: {brand}')

                        brand_url = self.get_brand_url(brand_code)
                        print(f'Brand URL: {brand_url}')
                        self.print_logs(f'Brand URL: {brand_url}')

                        auth_token = self.get_authorization_token()
                        self.print_logs(f'Authorization Token: {auth_token}')

                        
                        if auth_token:
                            headers = self.get_api_headers(auth_token)
                            self.print_logs(f'Headers: {headers}')
                            user_data = self.get_user_data(auth_token, store.username, headers)
                            self.print_logs(f'User Data: {user_data}')
                            brand_products_data = self.get_brand_products(brand_code, auth_token, headers, user_data)
                            for glasses_type in brand_with_type['glasses_type']:
                                start_time = datetime.now()

                                all_products = self.get_all_products_by_type(glasses_type, brand_products_data)
                                total_products = len(all_products)
                                scraped_products = 0

                                print(f'Type: {glasses_type} | Total products: {total_products}')
                                print(f'Start Time: {start_time.strftime("%A, %d %b %Y %I:%M:%S %p")}')

                                self.print_logs(f'Type: {glasses_type} | Total products: {total_products}')
                                self.print_logs(f'Start Time: {start_time.strftime("%A, %d %b %Y %I:%M:%S %p")}')

                                if total_products and int(total_products) > 0:
                                    self.printProgressBar(scraped_products, total_products, prefix = 'Progress:', suffix = 'Complete', length = 50)

                                for style_name in all_products:
                                    self.scrape_product(style_name, brand, glasses_type, auth_token, headers, user_data)
                                    self.save_to_json(self.data)
                                    scraped_products += 1
                                    if total_products and int(total_products) > 0:
                                        self.printProgressBar(scraped_products, total_products, prefix = 'Progress:', suffix = 'Complete', length = 50)
                                    
                                end_time = datetime.now()
                                print(f'End Time: {end_time.strftime("%A, %d %b %Y %I:%M:%S %p")}')
                                print('Duration: {}\n'.format(end_time - start_time))

                                self.print_logs(f'End Time: {end_time.strftime("%A, %d %b %Y %I:%M:%S %p")}')
                                self.print_logs('Duration: {}\n'.format(end_time - start_time))
                        else: 
                            print(f'Failed to get brand url for {brand}')
                            self.print_logs(f'Failed to get brand url for {brand}')

            else: 
                print(f'Failed to login \nURL: {store.link}\nUsername: {str(store.username)}\nPassword: {str(store.password)}')
                self.print_logs(f'Failed to login \nURL: {store.link}\nUsername: {str(store.username)}\nPassword: {str(store.password)}')
        except Exception as e:
            if self.DEBUG: print(f'Exception in MyMarchon_Scraper controller: {e}')
            self.print_logs(f'Exception in MyMarchon_Scraper controller: {e}')
        finally: 
            self.browser.quit()
            self.save_to_json(self.data)
    
    # initiate the browser
    def initiate_browser(self, store: Store):
        self.browser.get(store.link)
        self.wait_until_browsing()

    # wait until the browsing is completed
    def wait_until_browsing(self) -> None:
        while True:
            try:
                state = self.browser.execute_script('return document.readyState; ')
                if 'complete' == state: break
                else: sleep(0.2)
            except: pass

    def login(self, username: str, password: str) -> bool:
        login_flag = False
        try:
            while not login_flag:
                if self.wait_until_element_found(5, 'xpath', '//input[@id="username"]'):
                    self.input_credentials(username, password)
                    sleep(random.uniform(1, 5))
                else:
                    if self.wait_until_element_found(5, 'xpath', f'//span[contains(text(), "ACCT# {username}")]'):
                        if self.wait_until_element_found(5, 'xpath', '//img[@class="home-asset-image"]'):
                            login_flag = True
                            break
                        else:
                            self.browser.refresh()
                            sleep(10)
                            self.wait_until_browsing()
        except Exception as e:
            self.print_logs(f'Exception in login: {str(e)}')
            if self.DEBUG: print(f'Exception in login: {str(e)}')
        finally: return login_flag

    def input_credentials(self, username: str, password: str) -> None:
        try:
            sleep(random.uniform(1, 5))
            self.browser.find_element(By.XPATH, '//input[@id="username"]').send_keys(username)
            sleep(random.uniform(1, 5))
            self.browser.find_element(By.XPATH, '//input[@id="password"]').send_keys(password)
            sleep(random.uniform(1, 5))
            button = WebDriverWait(self.browser, 50).until(EC.element_to_be_clickable((By.XPATH, '//button[contains(@title, "Sign In")]')))
            sleep(random.uniform(1, 5))
            button.click()
            sleep(10)
            self.wait_until_browsing()
        except Exception as e:
            if self.DEBUG: print(f'Exception in input_credentials: {e}')
            self.print_logs(f'Exception in input_credentials: {e}')

    def wait_until_element_found(self, wait_value: int, type: str, value: str) -> bool:
        flag = False
        try:
            if type == 'id':
                WebDriverWait(self.browser, wait_value).until(EC.presence_of_element_located((By.ID, value)))
                flag = True
            elif type == 'xpath':
                WebDriverWait(self.browser, wait_value).until(EC.presence_of_element_located((By.XPATH, value)))
                flag = True
            elif type == 'css_selector':
                WebDriverWait(self.browser, wait_value).until(EC.presence_of_element_located((By.CSS_SELECTOR, value)))
                flag = True
            elif type == 'class_name':
                WebDriverWait(self.browser, wait_value).until(EC.presence_of_element_located((By.CLASS_NAME, value)))
                flag = True
            elif type == 'tag_name':
                WebDriverWait(self.browser, wait_value).until(EC.presence_of_element_located((By.TAG_NAME, value)))
                flag = True
        except: pass
        finally: return flag

    def get_authorization_token(self) -> str:
        authorization_token = ''
        try:
            for request in self.browser.requests:
                if request.response and 'authorization' in request.headers and 'Bearer' in request.headers.get('authorization'):
                    authorization_token = request.headers.get('authorization')
                    break
        except Exception as e:
            if self.DEBUG: print(f'Exception in get_authorization_token: {e}')
            self.print_logs(f'Exception in get_authorization_token: {e}')
        finally: return authorization_token

    def get_user_data(self, auth_token: str, username: str, headers: dict) -> dict:
        user_data = dict()
        try:
            API = 'https://api.mymarchon.com/OrganizationServicesWeb/getUserCredential2JS/invoke'
            json_data = {
                'userCredential': {
                    'token': auth_token,
                },
                'userID': username,
            }
            response = requests.post( url=API, headers=headers, json=json_data, verify=False)
            if response.status_code == 200: user_data = response.json().get('userCredential', {})
        except Exception as e:
            if self.DEBUG: print(f'Exception in get_user_data: {e}')
            self.print_logs(f'Exception in get_user_data: {e}')
        finally: return user_data

    def get_api_headers(self, auth_token: str) -> dict:
        return {
            'accept': 'application/json, text/plain, */*',
            'accept-language': 'en-US,en;q=0.9',
            'authorization': auth_token,
            'content-type': 'application/json',
            'origin': 'https://account.mymarchon.com',
            'priority': 'u=1, i',
            'referer': 'https://account.mymarchon.com/',
            'sec-ch-ua': '"Chromium";v="140", "Not=A?Brand";v="24", "Google Chrome";v="140"',
            'sec-ch-ua-mobile': '?0',
            'sec-ch-ua-platform': '"Windows"',
            'sec-fetch-dest': 'empty',
            'sec-fetch-mode': 'cors',
            'sec-fetch-site': 'same-site',
            'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/140.0.0.0 Safari/537.36',
        }
    
    def get_brand_url(self, brand_code: str) -> str:
        brand_url = ''
        try:
            brand_url = f'https://account.mymarchon.com/it/#/it/brand-collection/{brand_code}/eyewear'
            # doc_tree = html.fromstring(self.browser.page_source)
            # brand_url = str(doc_tree.xpath(f'//a[text()="{brand_name}"]/@href')[0]).strip()
        except Exception as e:
            if self.DEBUG: print(f'Exception in get_brand_url: {e}')
            self.print_logs((f'Exception in get_brand_url: {e}'))
        finally: return brand_url

    def open_new_tab(self, url: str) -> None:
        # open category in new tab
        self.browser.execute_script('window.open("'+str(url)+'","_blank");')
        self.browser.switch_to.window(self.browser.window_handles[len(self.browser.window_handles) - 1])
        self.wait_until_browsing()
    
    def close_last_tab(self) -> None:
        self.browser.close()
        self.browser.switch_to.window(self.browser.window_handles[len(self.browser.window_handles) - 1])

    def get_brand_products(self, brand_code: str, auth_token: str, headers: dict, user_data: dict):
        brand_products_data = dict()
        try:
            BRAND_API = 'https://api.mymarchon.com/ProductCatologWebWeb/catalog/catalog'
            json_data = {
                'userCredential': {
                    'userID': user_data.get('userID', ''),
                    'salesOrg': user_data.get('salesOrg', ''),
                    'defaultSalesOrg': user_data.get('salesOrg', ''),
                    'userType': user_data.get('userType', ''),
                    'name': user_data.get('name', ''),
                    'language': user_data.get('language', ''),
                    'phoneExtension': '',
                    'premierStatus': '',
                    'shipToNumber': '',
                    'accountNumber': user_data.get('accountNumber', ''),
                    'greenGrass': 'N',
                    'ftGreenGrass': 'N',
                    'currencyCode': 'EUR',
                    'sunRx': 'N',
                    'warrantyFeatureAvailable': True,
                    'buyingGroup': True,
                    'mktProgram': {
                        'programName': '',
                        'eligibility': '',
                        'optInStatus': 'NULL',
                        'autoShip': '',
                    },
                    'custSalesArea': [
                        {
                            'custno': user_data.get('accountNumber', ''),
                            'salesOrg': user_data.get('salesOrg', ''),
                            'salesOrgDesc': 'Marchon Italia',
                            'sapDivision': 99,
                            'sapDivisionDesc': 'Common Division',
                            'sapDistribution': 10,
                            'sapDistributionDesc': 'Optical Provider',
                        },
                    ],
                    'relatedSoldTos': [],
                    'token': auth_token,
                    'isFirstTimeGreenGrass': False,
                    'isGreenGrassAccount': False,
                    'isEnrolledInSunRx': False,
                    'isEnrolledInKaleyedoscope': False,
                    'isTestAccount': False,
                },
                'accountNumber': user_data.get('accountNumber', ''),
                'salesOrg': user_data.get('salesOrg', ''),
                'distChannel': 10,
                'soldTo': user_data.get('accountNumber', ''),
                'locale': user_data.get('language', ''),
                'brandCode': brand_code,
            }
            self.print_logs(f'API: {BRAND_API}')
            self.print_logs(f'Payload: {json_data}')
            response = requests.post(url=BRAND_API, headers=headers, json=json_data, verify=False)
            self.print_logs(f'Status Code: {response.status_code}')
            if response.status_code == 200: 
                brand_products_data = response.json()
                self.print_logs(f'Brand Products Data: {brand_products_data}')
        except Exception as e:
            if self.DEBUG: print(f'Exception in get_brand_data: {e}')
            self.print_logs(f'Exception in get_brand_data: {e}')
        finally: return brand_products_data
    
    def get_all_products_by_type(self, glasses_type: str, brand_products_data: dict) -> list[str]:
        all_products_numbers = []
        try:
            what_to_check = ''
            if glasses_type == 'Eyeglasses': what_to_check = ' Optical'
            elif glasses_type == 'Sunglasses': what_to_check = ' Sun'

            for catalogStyle in brand_products_data.get('catalog').get('catalogStyle'):
                if str(what_to_check).strip().lower() in str(catalogStyle.get('styleSkus')[0].get('marketingGroupDescription')).strip().lower():
                    if catalogStyle.get('style') not in all_products_numbers: 
                            all_products_numbers.append(catalogStyle.get('style'))
        except Exception as e:
            if self.DEBUG: print(f'Exception in get_all_products_by_type: {e}')
            self.print_logs(f'Exception in get_all_products_by_type: {e}')
        finally: return all_products_numbers

    def scrape_product(self, style_name: str, brand_name:str, glasses_type: str, auth_token: str, headers: dict, user_data: dict) -> None:
        try:
            PRODUCT_API = 'https://api.mymarchon.com/ProductCatologWebWeb/Frame/sku'
            json_data = {
                'userCredential': {
                    'userID': user_data.get('userID', ''),
                    'salesOrg': user_data.get('salesOrg', ''),
                    'defaultSalesOrg': user_data.get('salesOrg', ''),
                    'userType': user_data.get('userType', ''),
                    'name': user_data.get('name', ''),
                    'language': user_data.get('language', ''),
                    'phoneExtension': '',
                    'premierStatus': '',
                    'shipToNumber': '',
                    'accountNumber': user_data.get('accountNumber', ''),
                    'greenGrass': 'N',
                    'ftGreenGrass': 'N',
                    'currencyCode': 'EUR',
                    'sunRx': 'N',
                    'warrantyFeatureAvailable': True,
                    'buyingGroup': True,
                    'mktProgram': {
                        'programName': '',
                        'eligibility': '',
                        'optInStatus': 'NULL',
                        'autoShip': '',
                    },
                    'custSalesArea': [
                        {
                            'custno': user_data.get('accountNumber', ''),
                            'salesOrg': user_data.get('salesOrg', ''),
                            'salesOrgDesc': 'Marchon Italia',
                            'sapDivision': 99,
                            'sapDivisionDesc': 'Common Division',
                            'sapDistribution': 10,
                            'sapDistributionDesc': 'Optical Provider',
                        },
                    ],
                    'relatedSoldTos': [],
                    'token': auth_token,
                    'isFirstTimeGreenGrass': False,
                    'isGreenGrassAccount': False,
                    'isEnrolledInSunRx': False,
                    'isEnrolledInKaleyedoscope': False,
                    'isTestAccount': False,
                },
                'accountNumber': user_data.get('accountNumber', ''),
                'salesOrg': user_data.get('salesOrg', ''),
                'distChannel': 10,
                'currencyCode': 'EUR',
                'itemType': 'FRAME',
                'orderType': 'RX',
                'includeFrontAndTemples': 'X',
                'style': style_name,
            }
            self.print_logs(f'Product API: {PRODUCT_API}')
            self.print_logs(f'Payload: {json_data}')
            response = requests.post(url=PRODUCT_API,  headers=headers, json=json_data, verify=False)
            self.print_logs(f'Status Code: {response.status_code}')
            if response.status_code == 200:
                product_data = response.json()
                self.print_logs(f'Product Data for {style_name}: {product_data}')
                if 'skuDetail' in product_data:
                    frame_codes_with_sizes = self.get_all_frame_codes_and_sizes(product_data)
                    if frame_codes_with_sizes:
                        for frame_code_with_sizes in frame_codes_with_sizes:
                            product = Product()
                            product.brand = str(brand_name).strip().title()
                            product.type = glasses_type

                            product.frame_code = frame_code_with_sizes.get('frame_code')

                            for skuDetail in product_data.get('skuDetail'):
                                if skuDetail.get('itemType') == 'FRAME' and skuDetail.get('color') == product.frame_code:
                                    product.number = skuDetail.get('styleName') if 'styleName' in skuDetail else ''
                                    product.frame_color = skuDetail.get('familyColorDesc') if 'familyColorDesc' in skuDetail else ''

                                    metafields = Metafields()
                                    metafields.for_who = skuDetail.get('gender') if 'gender' in skuDetail else ''
                                    metafields.frame_material = skuDetail.get('planMaterial') if 'planMaterial' in skuDetail else ''
                                    metafields.img_url = skuDetail.get('colorImageURL') if 'colorImageURL' in skuDetail else ''
                                    if 'sku360Image' in skuDetail:
                                        metafields.img_360_urls = [sku360Image.get('image') for sku360Image in skuDetail.get('sku360Image')]
                                    product.metafields = metafields

                            for variant_list in frame_code_with_sizes.get('sizes'):
                                for skuDetail in product_data.get('skuDetail'):
                                    if skuDetail.get('itemType') == 'FRAME' and skuDetail.get('color') == product.frame_code and skuDetail.get('size') == variant_list:
                                        variant = Variant()
                                        variant.title = f'{product.number} {product.frame_code} {int(variant_list)}'
                                        variant.size = str(int(variant_list))
                                        variant.sku = str(variant.title).strip().replace(' ', '_')
                                        # variant.inventory_quantity = skuDetail.get('availableQty')
                                        variant.listing_price = str(skuDetail.get('msrp')).strip() if 'msrp' in skuDetail else ''
                                        variant.wholesale_price = str(skuDetail.get('retail')).strip() if 'retail' in skuDetail else ''
                                        variant.barcode_or_gtin = str(skuDetail.get('upcNumber')).strip() if 'upcNumber' in skuDetail else ''
                                        # variant.weight = skuDetail.get('weight')
                                        # variant.found_status = 1
                                        product.variants.append(variant)

                            self.data.append(product)
                    else:
                        self.print_logs(f'No frame codes found for {style_name}')
                else:
                    self.print_logs(f'No skuDetail found for {style_name}')
        except Exception as e:
            if self.DEBUG: print(f'Exception in scrape_product_data: {e}')
            self.print_logs(f'Exception in scrape_product_data: {e}')
    
    def get_all_frame_codes_and_sizes(self, product_data: dict) -> list[dict]:
        frame_codes_with_sizes = []
        try:
            frame_codes = []
            for skuDetail in product_data.get('skuDetail'):
                if skuDetail.get('color') not in frame_codes:
                    frame_codes.append(skuDetail.get('color'))
            
            for frame_code in frame_codes:
                variant_sizes = []
                for skuDetail in product_data.get('skuDetail'):
                    if frame_code == skuDetail.get('color') and skuDetail.get('itemType') == 'FRAME':
                        if 'size' in skuDetail and skuDetail.get('size') not in variant_sizes:
                            variant_sizes.append(skuDetail.get('size'))
                frame_codes_with_sizes.append({'frame_code': frame_code, 'sizes': variant_sizes})
        except Exception as e:
            if self.DEBUG: print(f'Exception in get_all_frame_codes_and_sizes: {e}')
            self.print_logs(f'Exception in get_all_frame_codes_and_sizes: {e}')
        finally: return frame_codes_with_sizes

    def save_to_json(self, products: list[Product]) -> None:
        try:
            json_products = []
            for product in products:
                json_product = {
                    'brand': product.brand, 
                    'number': product.number, 
                    'name': product.name, 
                    'frame_code': product.frame_code, 
                    'frame_color': product.frame_color, 
                    'lens_code': product.lens_code, 
                    'lens_color': product.lens_color, 
                    'status': product.status, 
                    'type': product.type, 
                    'url': product.url,
                    'metafields': {
                        'for_who': product.metafields.for_who,
                        'product_size': product.metafields.product_size,
                        'lens_material': product.metafields.lens_material,
                        'lens_technology': product.metafields.lens_technology,
                        'frame_material': product.metafields.frame_material,
                        'frame_shape': product.metafields.frame_shape,
                        'gtin1': product.metafields.gtin1,
                        'img_url': product.metafields.img_url,
                        'fitting_info': product.metafields.fitting_info,
                        'img_360_urls': product.metafields.img_360_urls
                    },
                    'variants': [
                        { 
                            'position': (index + 1), 
                            'title': variant.title, 
                            'sku': variant.sku, 
                            'inventory_quantity': variant.inventory_quantity, 
                            'found_status': variant.found_status, 
                            'listing_price': variant.listing_price, 
                            'wholesale_price': variant.wholesale_price, 
                            'barcode_or_gtin': variant.barcode_or_gtin, 
                            'size': variant.size, 
                            'weight': variant.weight 
                        } 
                        for index, variant in enumerate(product.variants)
                    ]
                }
                json_products.append(json_product)
            
           
            with open(self.result_filename, 'w') as f: json.dump(json_products, f)
            
        except Exception as e:
            if self.DEBUG: print(f'Exception in save_to_json: {e}')
            self.print_logs(f'Exception in save_to_json: {e}')
    
    # print logs to the log file
    def print_logs(self, log: str) -> None:
        try:
            with open(self.logs_filename, 'a') as f:
                f.write(f'\n{log}')
        except: pass

    def printProgressBar(self, iteration, total, prefix = '', suffix = '', decimals = 1, length = 100, fill = '█', printEnd = "\r") -> None:
        """
        Call in a loop to create terminal progress bar
        @params:
            iteration   - Required  : current iteration (Int)
            total       - Required  : total iterations (Int)
            prefix      - Optional  : prefix string (Str)
            suffix      - Optional  : suffix string (Str)
            decimals    - Optional  : positive number of decimals in percent complete (Int)
            length      - Optional  : character length of bar (Int)
            fill        - Optional  : bar fill character (Str)
            printEnd    - Optional  : end character (e.g. "\r", "\r\n") (Str)
        """
        percent = ("{0:." + str(decimals) + "f}").format(100 * (iteration / float(total)))
        filledLength = int(length * iteration // total)
        bar = fill * filledLength + '-' * (length - filledLength)
        print(f'\r{prefix} |{bar}| {percent}% {suffix}', end = printEnd)
        # Print New Line on Complete
        if iteration == total: 
            print()

def printProgressBar(iteration, total, prefix = '', suffix = '', decimals = 1, length = 100, fill = '█', printEnd = "\r") -> None:
        """
        Call in a loop to create terminal progress bar
        @params:
            iteration   - Required  : current iteration (Int)
            total       - Required  : total iterations (Int)
            prefix      - Optional  : prefix string (Str)
            suffix      - Optional  : suffix string (Str)
            decimals    - Optional  : positive number of decimals in percent complete (Int)
            length      - Optional  : character length of bar (Int)
            fill        - Optional  : bar fill character (Str)
            printEnd    - Optional  : end character (e.g. "\r", "\r\n") (Str)
        """
        percent = ("{0:." + str(decimals) + "f}").format(100 * (iteration / float(total)))
        filledLength = int(length * iteration // total)
        bar = fill * filledLength + '-' * (length - filledLength)
        print(f'\r{prefix} |{bar}| {percent}% {suffix}', end = printEnd)
        # Print New Line on Complete
        if iteration == total: 
            print()

def read_file(filename: str):
    f = open(filename)
    data = f.read()
    f.close()
    return data

def read_data_from_json_file(DEBUG, result_filename: str):
    data = []
    try:
        print('Downloading images...')
        
        files = glob.glob(result_filename)
        if files:
            json_data = json.loads(read_file(files[-1]))
            if json_data: printProgressBar(0, len(json_data), prefix = 'Progress:', suffix = 'Complete', length = 50)

            for index, json_d in enumerate(json_data):
                brand = json_d.get('brand') if 'brand' in json_d else ''
                
                number = str(json_d.get('number')).strip().upper() if 'number' in json_d else ''
                if '/' in number: number = number.replace('/', '-').strip()
                
                frame_code = str(json_d.get('frame_code')).strip().upper() if 'frame_code' in json_d else ''
                if '/' in frame_code: frame_code = frame_code.replace('/', '-').strip()

                frame_color = str(json_d.get('frame_color')).strip().title() if 'frame_color' in json_d else ''
                # lens_color = str(json_d['lens_color']).strip().title() if 'lens_color' in json_d else ''
                
                img_url = str(json_d.get('metafields').get('img_url')).strip() if 'img_url' in json_d.get('metafields') else ''
                # fitting_info = str(json_d.get('metafields').get('fitting_info')).strip() if 'fitting_info' in json_d.get('metafields') else ''
                for json_variant in json_d.get('variants', []):
                    sku = str(json_variant.get('sku')).strip().upper()
                    if '/' in sku: sku = sku.replace('/', '-').strip()
                    wholesale_price = str(json_variant.get('wholesale_price')).strip()
                    listing_price = str(json_variant.get('listing_price')).strip()
                    barcode_or_gtin = str(json_variant.get('barcode_or_gtin')).strip()

                    image_filename = f'Images/{sku}.jpg'
                    if not os.path.exists(image_filename): 
                        image_attachment = download_image(img_url)
                        if image_attachment:
                            with open(image_filename, 'wb') as f: f.write(image_attachment)
                            # crop_downloaded_image(f'Images/{sku}.jpg')
                    data.append([number, frame_code, frame_color, brand, sku, wholesale_price, listing_price, barcode_or_gtin])

                printProgressBar(index + 1, len(json_data), prefix = 'Progress:', suffix = 'Complete', length = 50)
    except Exception as e:
        if DEBUG: print(f'Exception in read_data_from_json_file: {e}')
        else: pass
    finally: return data

def download_image(url):
    image_attachment = ''
    try:
        headers = {
            'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
            'accept-Encoding': 'gzip, deflate, br',
            'accept-Language': 'en-US,en;q=0.9',
            'cache-Control': 'max-age=0',
            'sec-ch-ua': '"Google Chrome";v="95", "Chromium";v="95", ";Not A Brand";v="99"',
            'sec-ch-ua-mobile': '?0',
            'sec-ch-ua-platform': '"Windows"',
            'sec-fetch-dest': 'document',
            'sec-fetch-mode': 'navigate',
            'sec-fetch-site': 'none',
            'Sec-Fetch-User': '?1',
            'upgrade-insecure-requests': '1',
        }
        counter = 0
        while True:
            try:
                response = requests.get(url=url, headers=headers, timeout=20)
                # print(response.status_code)
                if response.status_code == 200:
                    # image_attachment = base64.b64encode(response.content)
                    image_attachment = response.content
                    break
                else: print(f'{response.status_code} found for downloading image')
            except: sleep(0.3)
            counter += 1
            if counter == 10: break
    except Exception as e: print(f'Exception in download_image: {str(e)}')
    finally: return image_attachment

def crop_downloaded_image(filename):
    try:
        im = Image.open(filename)
        width, height = im.size   # Get dimensions
        new_width = 1680
        new_height = 1020
        if width > new_width and height > new_height:
            left = (width - new_width)/2
            top = (height - new_height)/2
            right = (width + new_width)/2
            bottom = (height + new_height)/2
            im = im.crop((left, top, right, bottom))
            im.save(filename)
    except Exception as e: print(f'Exception in crop_downloaded_image: {e}')

def saving_picture_in_excel(data: list, excel_results_filename: str):
    workbook = Workbook()
    worksheet = workbook.active

    worksheet.cell(row=1, column=1, value='Model Code')
    worksheet.cell(row=1, column=2, value='Lens Code')
    worksheet.cell(row=1, column=3, value='Color Frame')
    worksheet.cell(row=1, column=4, value='Brand')
    worksheet.cell(row=1, column=5, value='SKU')
    worksheet.cell(row=1, column=6, value='Wholesale Price')
    worksheet.cell(row=1, column=7, value='Listing Price')
    worksheet.cell(row=1, column=8, value="UPC")
    worksheet.cell(row=1, column=9, value="Image")

    for index, d in enumerate(data):
        new_index = index + 2

        worksheet.cell(row=new_index, column=1, value=d[0])
        worksheet.cell(row=new_index, column=2, value=d[1])
        worksheet.cell(row=new_index, column=3, value=d[2])
        worksheet.cell(row=new_index, column=4, value=d[3])
        worksheet.cell(row=new_index, column=5, value=d[4])
        worksheet.cell(row=new_index, column=6, value=d[5])
        worksheet.cell(row=new_index, column=7, value=d[6])
        worksheet.cell(row=new_index, column=8, value=d[7])

        image = f'Images/{d[-4]}.jpg'
        if os.path.exists(image):
            im = Image.open(image)
            width, height = im.size
            worksheet.row_dimensions[new_index].height = height
            worksheet.add_image(Imag(image), anchor='I'+str(new_index))

    workbook.save(excel_results_filename)

DEBUG = True
try:
    pathofpyfolder = os.path.realpath(sys.argv[0])
    # get path of Exe folder
    path = pathofpyfolder.replace(pathofpyfolder.split('\\')[-1], '')
    if '.exe' in pathofpyfolder.split('\\')[-1]: DEBUG = False

    chrome_path = ChromeDriverManager().install()
    if 'chromedriver.exe' not in chrome_path:
        chrome_path = str(chrome_path).split('/')[0].strip()
        chrome_path = f'{chrome_path}\\chromedriver.exe'

    # create directories
    if not os.path.exists('requirements'): os.makedirs('requirements')
    if not os.path.exists('Logs'): os.makedirs('Logs')
    if not os.path.exists('Images'): os.makedirs('Images')
    
    start_json_filename = 'Start.json'
    credentails_filename = 'requirements/credentails.json'
    json_results_filename = 'requirements/json_results.json'
    excel_results_filename = 'Results.xlsx'
    
    # remove old files
    if os.path.exists(json_results_filename): os.remove(json_results_filename)
    if os.path.exists(excel_results_filename): os.remove(excel_results_filename)
    log_files = glob.glob('Logs/*.txt')
    if len(log_files) > 5:
        oldest_file = min(log_files, key=os.path.getctime)
        os.remove(oldest_file)
        log_files = glob.glob('Logs/*.txt')
    for filename in glob.glob('Images/*'): os.remove(filename)


    if os.path.exists(start_json_filename):
        json_data = json.loads(read_file(start_json_filename))
        
        if 'brands' in json_data:
            brands = json_data.get('brands')

            if os.path.exists(credentails_filename):
                credentails_json_data = json.loads(read_file(credentails_filename))

                store = Store()
                store.link = credentails_json_data.get('url')
                store.username = credentails_json_data.get('username')
                store.password = credentails_json_data.get('password')
                store.login_flag = True

                scrape_time = datetime.now().strftime('%d-%m-%Y %H-%M-%S')
                logs_filename = f'Logs/Logs {scrape_time}.txt'

                MyMarchon_Scraper(DEBUG, json_results_filename, logs_filename, chrome_path).controller(store, brands)
                
                data = read_data_from_json_file(DEBUG, json_results_filename)
                
                saving_picture_in_excel(data, excel_results_filename)

            else: print(f'No {credentails_filename} file found')
        else: print(f'No brands found in {start_json_filename} file')
    else: print(f'No {start_json_filename} file found')
    
except Exception as e: print('Exception: '+str(e))
